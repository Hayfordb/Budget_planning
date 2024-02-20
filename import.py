import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

class ExcelMergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Merger")
        self.geometry("600x400")

        self.file_treeview = ttk.Treeview(self, columns=("Files"), show="headings")
        self.file_treeview.heading("Files", text="Excel Files")
        self.file_treeview.pack(pady=20, fill=tk.BOTH, expand=True)

        select_folder_btn = tk.Button(self, text="Выберите папку", command=self.select_folder)
        select_folder_btn.pack(pady=10)

        merge_btn = tk.Button(self, text="Собрать", command=self.merge_excel_files)
        merge_btn.pack(pady=10)

        self.input_folder_path = ""
        self.files_to_merge = []

    def check_excel_parameters(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if "Параметры" in wb.sheetnames:
                sheet = wb["Параметры"]
                for row in sheet['C5:C13']:
                    for cell in row:
                        if cell.value is None or cell.value == "":
                            return False
                return True
            else:
                return False
        except InvalidFileException:
            return False  # Файл не может быть открыт как Excel файл.
        except Exception as e:
            print(f"Ошибка при проверке файла {file_path}: {e}")
            return False

    def select_folder(self):
        self.input_folder_path = filedialog.askdirectory(title="Выберите папку с файлами Excel")
        if not self.input_folder_path:
            return

        self.file_treeview.delete(*self.file_treeview.get_children())
        self.files_to_merge = []
        for file in os.listdir(self.input_folder_path):
            if file.endswith('.xlsx'):
                file_path = os.path.join(self.input_folder_path, file)
                valid = self.check_excel_parameters(file_path)
                if valid:
                    self.file_treeview.insert("", tk.END, values=(file), tags=('valid',))
                    self.files_to_merge.append(file_path)
                else:
                    self.file_treeview.insert("", tk.END, values=(file), tags=('invalid',))
        self.file_treeview.tag_configure('valid', foreground='black')
        self.file_treeview.tag_configure('invalid', foreground='red')

    def merge_excel_files(self):
        if not self.files_to_merge:
            messagebox.showerror("Ошибка", "Не выбраны файлы для слияния.")
            return

        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Сохранить конечный файл как")
        if not output_file_path:
            return

        all_transformed_data = []
        for file_path in self.files_to_merge:
            data = pd.read_excel(file_path, sheet_name="Расчёт")
            wb = openpyxl.load_workbook(file_path, data_only=True)
            parameters_sheet = wb["Параметры"]
            parameters = {}
            for row in parameters_sheet.iter_rows(min_row=5, max_row=13, min_col=2, max_col=3, values_only=True):
                parameter_name, parameter_value = row
                parameters[parameter_name] = parameter_value
            
            transformed_data_list = []
            for date_column in data.columns[1:]:
                if date_column == "Итого":
                    continue
                temp_df = data[['Статья', date_column]].copy()
                temp_df.rename(columns={date_column: 'План'}, inplace=True)
                temp_df['Дата-Статья'] = pd.to_datetime(date_column)
                # Добавление параметров в DataFrame
                for parameter_name in parameters:
                    temp_df[parameter_name] = parameters[parameter_name]
                transformed_data_list.append(temp_df)
            all_transformed_data.append(pd.concat(transformed_data_list, ignore_index=True))

        final_data = pd.concat(all_transformed_data, ignore_index=True)
        final_data.to_excel(output_file_path, index=False)
        messagebox.showinfo("Завершено", "Конечный файл успешно сохранен.")

if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()

